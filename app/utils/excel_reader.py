import json
import re

import httpx
from openpyxl import load_workbook


def valid_uuid(uuid: str | None):
    regex = re.compile(r'^[a-f0-9]{8}-?[a-f0-9]{4}-?4[a-f0-9]{3}-?[89ab][a-f0-9]{3}-?[a-f0-9]{12}\Z', re.I)
    if uuid:
        match = regex.match(uuid)
        return bool(match)
    return False


async def convert_excel_to_json():
    admin_json = []
    book = load_workbook('app/admin/Menu.xlsx')
    sheet = book['Лист1']
    for row in range(1, sheet.max_row + 1):
        id = sheet[f'A{row}'].value
        if valid_uuid(id):
            title = sheet[f'B{row}'].value
            description = sheet[f'C{row}'].value
            admin_json.append({'id': id, 'title': title, 'description': description, 'submenus': []})
            for row in range(row, sheet.max_row + 1):
                id = sheet[f'B{row+1}'].value
                if type(id) == str and valid_uuid(id) is False:
                    break
                if valid_uuid(id):
                    title = sheet[f'C{row+1}'].value
                    description = sheet[f'D{row + 1}'].value
                    admin_json[-1]['submenus'].append({'id': id, 'title': title,
                                                      'description': description, 'dishes': []})
                    for row in range(row, sheet.max_row + 1):
                        id = sheet[f'C{row+2}'].value
                        if type(id) == str and valid_uuid(id) is False:
                            break
                        if valid_uuid(id):
                            title = sheet[f'D{row+2}'].value
                            description = sheet[f'E{row+2}'].value
                            price = str(sheet[f'F{row+2}'].value)
                            admin_json[-1]['submenus'][-1]['dishes'].append(
                                {'id': id, 'title': title, 'description': description, 'price': price})
    admin_json = json.dumps(admin_json, ensure_ascii=False)
    return admin_json


async def compare():
    menu = httpx.get('http://127.0.0.1:8000/api/v1/menus/all_menu')
    print('GET', menu)
